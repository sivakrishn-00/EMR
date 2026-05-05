from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
from reportlab.lib.units import inch
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from patients.report_theme import (
    NAVY, get_report_styles, create_main_bar, create_sub_header, 
    get_standard_table_style, draw_elite_border
)

def PageBorder(canvas, doc):
    draw_elite_border(canvas, doc, 
        footer_left="PHARMACY AUDIT & COMPLIANCE", 
        footer_center="Hospital Grade Consumption Audit | Powered by Bavya Health Service PVT Ltd.")

def generate_consumption_pdf_report(data):
    """
    Elite Admin Pass: Implementing the Pharmacy Consumption Audit Report using the 
    Hybrid Elite Design System to eliminate code redundancy and maintain brand parity.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=30, bottomMargin=40)
    
    theme = get_report_styles()
    std_grid = get_standard_table_style()
    
    elements = []

    # --- 1. HEADER ---
    header_data = [[
        Paragraph(f"<font color='#1e3a8a' size=15><b>ADMINISTRATIVE AUDIT</b></font><br/><font size=7 color='#64748b'>Pharmacy & Inventory Intelligence<br/>Project: {data.get('project_name', 'GLOBAL')}</font>", theme['value']),
        Paragraph("<font size=18 color='#1e3a8a'><b>CONSUMPTION REPORT</b></font><br/><font size=7 color='#94a3b8'>PHARMACY DEPARTMENT</font>", theme['badge_green'])
    ]]
    elements.append(Table(header_data, colWidths=[3.1*inch, 3.5*inch]))
    elements.append(Spacer(1, 0.1 * inch))
    elements.append(Table([[""]], colWidths=[6.6*inch], rowHeights=[3], style=[('BACKGROUND', (0,0), (-1,-1), NAVY)]))
    elements.append(Spacer(1, 0.2 * inch))

    # --- 2. SUMMARY (KEEP SOLID BAR) ---
    elements.append(create_main_bar("CONSUMPTION SUMMARY"))
    elements.append(Spacer(1, 0.05 * inch))
    
    summary_data = [
        [Paragraph("TOTAL VISITS", theme['label_left']), Paragraph(str(data.get('total_visits', 0)), theme['value']), Paragraph("TOTAL PATIENTS", theme['label_left']), Paragraph(str(data.get('total_patients', 0)), theme['value'])],
        [Paragraph("GRAND TOTAL UNITS", theme['label_left']), Paragraph(str(data.get('grand_total_units', 0)), theme['value']), Paragraph("GRAND TOTAL COST", theme['label_left']), Paragraph(f"₹{data.get('grand_total_cost', 0):,.2f}", theme['value'])],
    ]
    sum_t = Table(summary_data, colWidths=[1.4*inch, 1.9*inch, 1.4*inch, 1.9*inch])
    sum_t.setStyle(std_grid)
    elements.append(sum_t)
    elements.append(Spacer(1, 0.3 * inch))

    # --- 3. DETAILED LOG (HYBRID MINIMALIST) ---
    elements.append(create_sub_header("DETAILED CONSUMPTION LOG"))
    elements.append(Spacer(1, 0.06 * inch))
    
    log_data = [[
        Paragraph("DATE", theme['label_left']), 
        Paragraph("PATIENT ID / NAME", theme['label_left']), 
        Paragraph("MEDICATION", theme['label_left']), 
        Paragraph("QTY", theme['label_center']), 
        Paragraph("COST", theme['label_center'])
    ]]
    
    for visit in data.get('items', []):
        for med in visit.get('medications', []):
            log_data.append([
                Paragraph(visit['visit_date'], theme['value']),
                Paragraph(f"<b>{visit['patient_id']}</b><br/>{visit['patient_name']}", theme['value']),
                Paragraph(med['name'], theme['value']),
                Paragraph(str(med['quantity']), theme['badge_green']),
                Paragraph(f"₹{med['total_cost']:,.2f}", theme['value'])
            ])

    log_t = Table(log_data, colWidths=[1.0*inch, 1.8*inch, 1.8*inch, 0.8*inch, 1.2*inch])
    log_t.setStyle(std_grid)
    elements.append(log_t)

    doc.build(elements, onFirstPage=PageBorder, onLaterPages=PageBorder)
    buffer.seek(0)
    return buffer

def generate_consumption_xlsx_report(data):
    """
    Elite Audit Pass: Generates a high-fidelity XLSX workbook for medication consumption audit.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumption Audit"
    
    # --- STYLES ---
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    sub_header_fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
    sub_header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin', color='e2e8f0'), 
                   right=Side(style='thin', color='e2e8f0'), 
                   top=Side(style='thin', color='e2e8f0'), 
                   bottom=Side(style='thin', color='e2e8f0'))

    # --- 1. TITLE & METADATA ---
    ws.merge_cells('A1:H1')
    ws['A1'] = "PHARMACY CONSUMPTION AUDIT REPORT"
    ws['A1'].font = Font(size=18, bold=True, color="1e3a8a")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws['A2'] = "Project Name:"
    ws['B2'] = data.get('project_name', 'GLOBAL')
    ws['A3'] = "Audit Period:"
    range_str = f"{data.get('start_date', 'Start')} to {data.get('end_date', 'End')}" if data.get('start_date') else f"Range: {data.get('range', 'N/A')}"
    ws['B3'] = range_str
    ws['G2'] = "Generated At:"
    ws['H2'] = data.get('generated_at', 'N/A')
    
    # Apply some styling to metadata labels
    for cell in [ws['A2'], ws['A3'], ws['G2']]:
        cell.font = Font(bold=True, color="64748b")

    # --- 2. SUMMARY DASHBOARD (Professional Bar) ---
    ws.append([]) # Spacer
    ws.append(["EXECUTIVE SUMMARY"])
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')
    ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    ws.append(["Metric", "Total Count", "", "", "Financial Overview", "Total Value"])
    ws.append(["Total Patient Visits", data.get('total_visits', 0), "", "", "Grand Total Cost", f"₹{data.get('grand_total_cost', 0):,.2f}"])
    ws.append(["Unique Patients Audited", data.get('total_patients', 0), "", "", "Total Units Dispensed", data.get('grand_total_units', 0)])
    
    # Style the mini-summary
    for row_idx in range(ws.max_row-2, ws.max_row+1):
        for col_idx in [1, 5]:
            ws.cell(row=row_idx, column=col_idx).font = Font(bold=True, color="1e3a8a")
    
    # --- 3. MAIN DATA TABLE ---
    ws.append([]) # Spacer
    ws.append([]) # Spacer
    headers = ["Date", "Card No", "Patient ID", "Patient Name", "Medication Name", "Quantity", "Unit Cost", "Total Cost"]
    ws.append(headers)
    
    # Style Headers
    header_row = ws.max_row
    for col in range(1, 9):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add Data
    for visit in data.get('items', []):
        for med in visit.get('medications', []):
            row = [
                visit['visit_date'],
                visit['card_no'],
                visit['patient_id'],
                visit['patient_name'],
                med['name'],
                med['quantity'],
                med['unit_price'],
                med['total_cost']
            ]
            ws.append(row)
            
            # Add borders and alternating row colors (subtle)
            for col in range(1, 9):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = border
                cell.alignment = center_align
                if ws.max_row % 2 == 0:
                    cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

    # --- 4. VISIT-WISE COLORFUL SUMMARY (MNC STYLE) ---
    ws.append([]) # Spacer
    ws.append([]) # Spacer
    ws.append(["VISIT-WISE CONSUMPTION ANALYTICS"])
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')
    summary_header_cell = ws.cell(row=ws.max_row, column=1)
    summary_header_cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid") # Emerald Green
    summary_header_cell.font = Font(color="FFFFFF", bold=True, size=12)
    summary_header_cell.alignment = center_align

    for visit in data.get('items', []):
        ws.append([f"VISIT DATE: {visit['visit_date']}", f"PATIENT: {visit['patient_name']} ({visit['card_no']})", "", "", "", "", "TOTAL UNITS:", visit['total_visit_units']])
        # Style visit header row
        curr_row = ws.max_row
        ws.merge_cells(f'B{curr_row}:F{curr_row}')
        for col in range(1, 9):
            cell = ws.cell(row=curr_row, column=col)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="ecfdf5", end_color="ecfdf5", fill_type="solid")
            cell.border = border
            cell.alignment = center_align

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_length = 0
        column_cell = col[0]
        column_letter = get_column_letter(column_cell.column)
        
        for cell in col:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40) # Cap at 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
